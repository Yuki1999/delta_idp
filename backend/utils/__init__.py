"""XLSX document parsing and conversion utilities."""

import io
import os
import base64
import tempfile
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import json


def parse_xlsx_to_structured_text(file_path: str) -> Dict[str, Any]:
    """
    Parse an XLSX file and extract structured text content.
    Handles merged cells and complex layouts.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {
        "sheets": [],
        "filename": os.path.basename(file_path),
    }

    for ws in wb.worksheets:
        sheet_data = {
            "name": ws.title,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "merged_cells": [str(mc) for mc in ws.merged_cells.ranges],
            "cells": [],
        }

        # Collect all non-empty cells with their positions
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    sheet_data["cells"].append({
                        "row": cell.row,
                        "col": cell.column,
                        "coordinate": cell.coordinate,
                        "value": str(cell.value).strip(),
                        "col_letter": get_column_letter(cell.column),
                    })

        result["sheets"].append(sheet_data)

    wb.close()
    return result


def xlsx_to_markdown_table(file_path: str) -> str:
    """
    Convert an XLSX file to a markdown representation.
    Best-effort conversion preserving spatial layout.
    """
    data = parse_xlsx_to_structured_text(file_path)
    md_parts = [f"# Document: {data['filename']}\n"]

    for sheet in data["sheets"]:
        md_parts.append(f"## Sheet: {sheet['name']}\n")

        # Organize cells by row
        rows: Dict[int, Dict[int, str]] = {}
        max_col = 0
        for cell in sheet["cells"]:
            r, c = cell["row"], cell["col"]
            if r not in rows:
                rows[r] = {}
            rows[r][c] = cell["value"]
            max_col = max(max_col, c)

        # Generate markdown
        for row_num in sorted(rows.keys()):
            row_cells = rows[row_num]
            line_parts = []
            for col in range(1, max_col + 1):
                val = row_cells.get(col, "")
                if val:
                    line_parts.append(f"**[C{col}]** {val}")
            if line_parts:
                md_parts.append(f"- Row {row_num}: " + " | ".join(line_parts))

        md_parts.append("")

    return "\n".join(md_parts)


def xlsx_to_plain_text(file_path: str) -> str:
    """
    Extract plain text from XLSX, preserving logical reading order.
    Groups nearby cells to reconstruct form fields.
    """
    data = parse_xlsx_to_structured_text(file_path)
    text_parts = []

    for sheet in data["sheets"]:
        text_parts.append(f"--- Sheet: {sheet['name']} ---")

        # Group cells by row
        rows: Dict[int, List[Dict]] = {}
        for cell in sheet["cells"]:
            r = cell["row"]
            if r not in rows:
                rows[r] = []
            rows[r].append(cell)

        for row_num in sorted(rows.keys()):
            row_cells = sorted(rows[row_num], key=lambda c: c["col"])
            # Group cells that are close together (within 5 columns)
            line = ""
            last_col = -100
            for cell in row_cells:
                if cell["col"] - last_col <= 5:
                    line += " " + cell["value"]
                else:
                    if line:
                        text_parts.append(line.strip())
                    line = cell["value"]
                last_col = cell["col"]
            if line:
                text_parts.append(line.strip())

    return "\n".join(text_parts)


def xlsx_to_images(file_path: str, output_dir: str, max_size: int = 1600) -> List[str]:
    """
    Render each sheet of an XLSX as an image.
    Returns list of image paths.
    This is a simplified renderer - uses openpyxl data with PIL.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    image_paths = []

    base_name = Path(file_path).stem

    for ws_idx, ws in enumerate(wb.worksheets):
        # Collect cell data for rendering
        cells = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200),
                                 max_col=min(ws.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    cells.append({
                        "row": cell.row,
                        "col": cell.column,
                        "value": str(cell.value),
                    })

        if not cells:
            continue

        # Calculate dimensions
        font_size = 12
        col_width = 120
        row_height = 22
        padding = 10

        max_row = max(c["row"] for c in cells)
        max_col = max(c["col"] for c in cells)

        img_width = (max_col + 1) * col_width + padding * 2
        img_height = (max_row + 1) * row_height + padding * 2

        # Limit image size
        scale = 1.0
        if img_width > max_size or img_height > max_size:
            scale = min(max_size / img_width, max_size / img_height)

        img = Image.new("RGB", (int(img_width * scale), int(img_height * scale)), "white")
        draw = ImageDraw.Draw(img)

        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                                       max(8, int(font_size * scale)))
        except (OSError, IOError):
            font = ImageFont.load_default()

        for cell in cells:
            x = (cell["col"]) * col_width * scale + padding * scale
            y = (cell["row"] - 1) * row_height * scale + padding * scale
            # Truncate long values
            val = cell["value"][:60]
            draw.text((x, y), val, fill="black", font=font)

        img_path = os.path.join(output_dir, f"{base_name}_sheet{ws_idx}.png")
        img.save(img_path, "PNG")
        image_paths.append(img_path)

    wb.close()
    return image_paths


def xlsx_to_image_base64(file_path: str) -> str:
    """
    Convert first sheet of XLSX to a base64-encoded PNG image.
    For sending to vision models.
    """
    import tempfile
    with tempfile.TemporaryDirectory() as tmpdir:
        paths = xlsx_to_images(file_path, tmpdir)
        if paths:
            with open(paths[0], "rb") as f:
                return base64.b64encode(f.read()).decode("utf-8")
        return ""


def xlsx_to_pdf(file_path: str, output_path: str = "") -> str:
    """
    Convert an XLSX file to PDF using LibreOffice headless.
    Preserves full layout, formatting, merged cells, fonts, etc.
    Returns the path to the generated PDF file.
    """
    import subprocess
    import shutil

    abs_path = os.path.abspath(file_path)
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"Source file not found: {abs_path}")

    if not output_path:
        base = Path(abs_path).stem
        output_path = os.path.join(tempfile.gettempdir(), f"{base}.pdf")

    # LibreOffice outputs to a directory with the same stem name + .pdf
    out_dir = tempfile.mkdtemp(prefix="lo_pdf_")
    try:
        cmd = [
            "libreoffice",
            "--headless",
            "--calc",
            "--convert-to", "pdf",
            "--outdir", out_dir,
            abs_path,
        ]
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=120
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice conversion failed (code {result.returncode}): "
                f"{result.stderr or result.stdout}"
            )

        # Find the generated PDF in output dir
        generated_name = Path(abs_path).stem + ".pdf"
        generated_path = os.path.join(out_dir, generated_name)
        if not os.path.exists(generated_path):
            # Fallback: look for any .pdf in the output dir
            pdfs = [f for f in os.listdir(out_dir) if f.endswith(".pdf")]
            if pdfs:
                generated_path = os.path.join(out_dir, pdfs[0])
            else:
                raise RuntimeError(f"No PDF generated. LibreOffice output: {result.stdout}")

        # Move to the desired output path
        shutil.move(generated_path, output_path)
        print(f"[xlsx_to_pdf] LibreOffice converted: {output_path} ({os.path.getsize(output_path)} bytes)")
        return output_path
    finally:
        # Cleanup temp dir
        shutil.rmtree(out_dir, ignore_errors=True)


def extract_key_value_pairs(file_path: str) -> Dict[str, str]:
    """
    Try to extract key-value pairs from the XLSX form layout.
    Looks for patterns like "Label: Value" or label in one cell, value in adjacent.
    """
    data = parse_xlsx_to_structured_text(file_path)
    kv_pairs = {}

    known_labels = [
        "Shipper/Export", "No. & date of invoice", "invoice", "No. & date of L/C",
        "Port of loading", "Final destination", "Carrier", "Sailing on or about",
        "For Account & Risk of Messrs", "Notify party", "L/C Issuing Bank",
        "Remarks", "Marks and numbers", "Description of good",
        "Quantity/Unit", "Unit Price", "Amount",
        "Net Weight", "Gross Weight", "Measurement",
        "P/O", "P/N", "SAMSUNG P/N", "ITEM",
    ]

    for sheet in data["sheets"]:
        # Build row-column map
        grid: Dict[Tuple[int, int], str] = {}
        for cell in sheet["cells"]:
            grid[(cell["row"], cell["col"])] = cell["value"]

        # Try to find labels and their values (same row, nearby column)
        for (r, c), val in grid.items():
            for label in known_labels:
                if label.lower() in val.lower():
                    # Check next cells to the right for a value
                    for offset in range(1, 10):
                        next_val = grid.get((r, c + offset), "")
                        if next_val and not any(
                            lab.lower() in next_val.lower() for lab in known_labels
                        ):
                            kv_pairs[label] = next_val
                            break
                    break

    return kv_pairs
