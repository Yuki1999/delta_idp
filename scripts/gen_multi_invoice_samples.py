"""
基于真实的 DS12650253_IV(1).xlsx / _PL(1).xlsx 模板生成 3 份"分批合并申报"测试样本。

策略：加载原始 xlsx，只改动这些单元格的值（保留全部样式/布局/合并单元格）：
- IV：BD7（发票号+日期）、G52（页脚发票号）、明细行 r28~r45 + r57~r74、合计行 r75
- PL：BB7（发票号+日期）、D52/D110（页脚发票号）、明细行 r26~r44 + r57~r102 + r115~r132、合计行 r133
- 明细：只写前 N 行、其余置空（保留原格线/样式）
- 合计：写死数字（不留公式，避免 Excel 重算差异）
"""
import os
import shutil
from openpyxl import load_workbook

SRC_IV = "/home/qqr/delta_idp/samples/DS12650253/DS12650253_IV(1).xlsx"
SRC_PL = "/home/qqr/delta_idp/samples/DS12650253/DS12650253_PL(1).xlsx"
OUT_DIR = "/home/qqr/delta_idp/samples/DS12650253_MULTI"
os.makedirs(OUT_DIR, exist_ok=True)

INV_DATE = "MAY.11,2026"

# ── 三份 part 的明细数据（沿用之前设计好的一套） ─────────────────
PARTS = [
    {
        "invoice_no": "DS12650253-1",
        "items": [
            {"item": 5,  "po": "XHD01-20251009-015", "spn": "CL10B332KB8WPNC", "qty": 164000, "up": 1},
            {"item": 8,  "po": "XHD01-20251027-003", "spn": "CL10B104KB8WPNC", "qty": 668000, "up": 2},
            {"item": 22, "po": "XHD01-20251027-003", "spn": "CL21B474KBFVPNE", "qty":  44000, "up": 3},
            {"item": 26, "po": "XHD01-20251027-003", "spn": "CL32Y106KBJ4PNE", "qty": 202000, "up": 4},
        ],
    },
    {
        "invoice_no": "DS12650253-2",
        "items": [
            {"item": 8,  "po": "XHD01-20251105-027", "spn": "CL31B106KOHVPNE", "qty": 224000, "up": 1},
            {"item": 7,  "po": "XHD01-20251110-020", "spn": "CL21B104KBFWPNE", "qty":  80000, "up": 2},
            {"item": 10, "po": "XHD01-20251110-020", "spn": "CL21Y475KBBVPNE", "qty": 260000, "up": 3},
            {"item": 2,  "po": "XHD01-20260106-035", "spn": "CL10B103KC8WPNC", "qty":  36000, "up": 4},
        ],
    },
    {
        "invoice_no": "DS12650253-3",
        "items": [
            {"item": 1,  "po": "XHD01-20260126-008", "spn": "CL10B103KC8WPNC", "qty":  32000, "up": 1},
            {"item": 2,  "po": "XHD01-20260130-026", "spn": "CL10B103KC8WPNC", "qty": 188000, "up": 2},
            {"item": 9,  "po": "XHD01-20260213-009", "spn": "CL05B472KB5VPNC", "qty":  90000, "up": 3},
            {"item": 14, "po": "XHD01-20260213-009", "spn": "CL05C470JB51PNC", "qty":  50000, "up": 4},
        ],
    },
]

PER_CTN_NET_KG = 5.900
PER_CTN_GROSS_KG = 6.500
PER_CTN_VOL_CBM = 0.027


def _fmt_int(n): return f"{n:,}"
def _fmt_kg(n): return f"{n:.3f}KG"
def _fmt_kgs(n): return f"{n:.3f}KGS"
def _fmt_cbm(n): return f"{n:.3f}CBM"


# 明细在原 IV 里可用的行号：r28~r45 (18 行) + r57~r74 (18 行) = 36 个位置
IV_DETAIL_ROWS = list(range(28, 46)) + list(range(57, 75))
# PL 里：r26~r44 (19) + r57~r102 (46) + r115~r132 (18) = 83 个位置
PL_DETAIL_ROWS = list(range(26, 45)) + list(range(57, 103)) + list(range(115, 133))

# 单元格里之前非 None 的列（清空时要覆盖到）：
# IV 明细行涉及的列（从原始 dump 看到）：B, L, AC(有时), AT, BQ, CE, CT
IV_DETAIL_COLS = ["B", "L", "AC", "AT", "BQ", "CE", "CT"]
# PL 明细行涉及的列：D, R, Z, AO(有时), BH, CQ, DB
PL_DETAIL_COLS = ["D", "R", "Z", "AO", "BH", "CQ", "DB"]


def build_invoice(part, out_path):
    shutil.copy(SRC_IV, out_path)
    wb = load_workbook(out_path)
    ws = wb.active  # LEIV001

    # 1) 顶部发票号（BD7），保留原来那种"发票号 + 空格 + 日期"的写法
    ws["BD7"] = f"{part['invoice_no']}               {INV_DATE}"
    # 2) 页脚发票号（G52）
    ws["G52"] = part["invoice_no"]

    # 3) 明细：只写前 N 行，其余全部清空（保留边框样式）
    items = part["items"]
    for idx, row_r in enumerate(IV_DETAIL_ROWS):
        if idx < len(items):
            it = items[idx]
            ws.cell(row=row_r, column=2).value  = it["item"]          # B
            ws.cell(row=row_r, column=12).value = it["po"]            # L
            ws.cell(row=row_r, column=29).value = None                # AC (P/N 未使用)
            ws.cell(row=row_r, column=46).value = it["spn"]           # AT (SAMSUNG P/N)
            ws.cell(row=row_r, column=69).value = _fmt_int(it["qty"]) # BQ (数量)
            ws.cell(row=row_r, column=83).value = it["up"]            # CE (单价)
            amount = it["qty"] * it["up"] / 1000
            ws.cell(row=row_r, column=98).value = amount              # CT (金额, 写死值)
        else:
            # 清空
            for col in IV_DETAIL_COLS:
                ws[f"{col}{row_r}"].value = None

    # 4) 合计行（r75）
    total_qty = sum(it["qty"] for it in items)
    total_amt = sum(it["qty"] * it["up"] / 1000 for it in items)
    ws["G75"] = "TOTAL :"
    ws["BL75"] = _fmt_int(total_qty)   # 数量合计
    ws["CO75"] = round(total_amt, 2)   # 金额合计（覆盖原公式）

    wb.save(out_path)


def build_packing_list(part, out_path):
    shutil.copy(SRC_PL, out_path)
    wb = load_workbook(out_path)
    ws = wb.active  # LEPL001

    # 1) 顶部发票号（BB7）
    ws["BB7"] = f"{part['invoice_no']}               {INV_DATE}"
    # 2) 页脚发票号（D52 页1、D110 页2）
    ws["D52"] = part["invoice_no"]
    ws["D110"] = part["invoice_no"]

    # 3) 明细：每行 1 CTN
    items = part["items"]
    for idx, row_r in enumerate(PL_DETAIL_ROWS):
        if idx < len(items):
            it = items[idx]
            ctn_no = idx + 1
            ws.cell(row=row_r, column=4).value  = ctn_no              # D  (C/T NO)
            ws.cell(row=row_r, column=18).value = it["item"]          # R  (ITEM)
            ws.cell(row=row_r, column=26).value = it["po"]            # Z  (P/O No)
            ws.cell(row=row_r, column=41).value = None                # AO (PART No 未用)
            ws.cell(row=row_r, column=60).value = it["spn"]           # BH (SAMSUNG P/N)
            ws.cell(row=row_r, column=95).value = _fmt_int(it["qty"]) # CQ (数量)
            ws.cell(row=row_r, column=106).value = f"{_fmt_int(it['qty'])} x0001"  # DB (拆箱明细)
        else:
            for col in PL_DETAIL_COLS:
                ws[f"{col}{row_r}"].value = None

    # 4) 合计行（r133）
    ctn_count = len(items)
    total_qty = sum(it["qty"] for it in items)
    total_net = ctn_count * PER_CTN_NET_KG
    total_gross = ctn_count * PER_CTN_GROSS_KG
    total_vol = ctn_count * PER_CTN_VOL_CBM
    ws["D133"] = "TOTAL"
    ws["X133"] = f"{ctn_count}C/T"
    ws["AN133"] = f"{_fmt_int(total_qty)}PC"
    ws["BI133"] = _fmt_kg(total_net)
    ws["CG133"] = _fmt_kgs(total_gross)
    ws["DA133"] = _fmt_cbm(total_vol)

    wb.save(out_path)


def main():
    print(f"输出目录：{OUT_DIR}")
    # 先清掉旧的简易版本
    for f in os.listdir(OUT_DIR):
        if f.endswith(".xlsx"):
            os.remove(os.path.join(OUT_DIR, f))
            print(f"  删除旧文件：{f}")

    for part in PARTS:
        inv_out = os.path.join(OUT_DIR, f"{part['invoice_no']}_IV.xlsx")
        pl_out  = os.path.join(OUT_DIR, f"{part['invoice_no']}_PL.xlsx")
        build_invoice(part, inv_out)
        build_packing_list(part, pl_out)
        print(f"  生成：{os.path.basename(inv_out)}  +  {os.path.basename(pl_out)}")

    # 汇总预期
    total_qty = sum(sum(x["qty"] for x in p["items"]) for p in PARTS)
    total_amt = sum(sum(x["qty"] * x["up"] / 1000 for x in p["items"]) for p in PARTS)
    total_ctn = sum(len(p["items"]) for p in PARTS)
    print("\n合并预期：")
    print(f"  发票号:  {' / '.join(p['invoice_no'] for p in PARTS)}")
    print(f"  数量:    {total_qty:>10,} PC")
    print(f"  金额:    {total_amt:>10,.2f} RMB")
    print(f"  箱数:    {total_ctn:>10d} C/T")
    print(f"  净重:    {total_ctn * PER_CTN_NET_KG:>10.3f} KG")
    print(f"  毛重:    {total_ctn * PER_CTN_GROSS_KG:>10.3f} KG")
    print(f"  体积:    {total_ctn * PER_CTN_VOL_CBM:>10.4f} CBM")


if __name__ == "__main__":
    main()
